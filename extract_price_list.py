# -*- coding: utf-8 -*-
"""ดึงตารางราคาข้อมูลดาวเทียมจาก Gistda_Price_List.pdf
แล้วส่งออกเป็น CSV / XLSX / HTML (ตารางเดียวรวมทุกหมวด)

ที่มา: https://www.gistda.or.th/download/Gistda_Price_List.pdf
"""
import re
from pathlib import Path

import pdfplumber
import pandas as pd

BASE = Path(__file__).parent
PDF = BASE / "Gistda_Price_List.pdf"
STEM = "Gistda_Price_List"

VAT_NOTE = "ราคาดังกล่าวยังไม่รวมภาษีมูลค่าเพิ่ม"
BULLET = ""  # อักขระ bullet ใน PDF ใช้ระบุแถวหมายเหตุ

COLUMNS = [
    "หมวด",
    "ดาวเทียม (Satellite)",
    "โหมด (Mode)",
    "รายละเอียดภาพ (Resolution)",
    "Polarization",
    "ข้อมูลในคลัง (Standard Archive)",
    "ข้อมูลชนิดสั่งถ่าย (Standard Tasking)",
    "Single Look Complex",
    "Path Image",
    "New Acquisition",
    "การติดตาม (Monitoring)",
    "หน่วย",
    "หมายเหตุ",
]
PRICE_COLUMNS = COLUMNS[5:11]
ARCHIVE, TASKING, SLC, PATH_IMG, NEW_ACQ, MONITOR = PRICE_COLUMNS

# ---------------------------------------------------------------- text clean
GARBLE = {
    "มลู คา่ เพิ่ม": "มูลค่าเพิ่ม",
    "ทปี่ รับ": "ที่ปรับ",
    "ขั้นตา่": "ขั้นต่ำ",
    "ป ี": "ปี",
}


def clean(text):
    """ล้างข้อความจาก PDF: bullet, สระอำแตก, คำเพี้ยน"""
    if text is None:
        return ""
    t = text.replace(BULLET, " ").replace("\n", " ")
    # สระอำที่ PDF แตกเป็น <พยัญชนะ/วรรณยุกต์> + ช่องว่าง + า
    t = re.sub(r"([ก-๎]) า", r"\1ำ", t)
    for bad, good in GARBLE.items():
        t = t.replace(bad, good)
    return re.sub(r"\s+", " ", t).strip()


def is_price(cell):
    return bool(re.fullmatch(r"[\d,]+", cell)) or cell == "N/A"


def price(cell):
    return None if cell in ("", "N/A") else int(cell.replace(",", ""))


def compact(row):
    """คืนเฉพาะเซลล์ที่ไม่ว่าง (ตัดข้อความซ้ำติดกันจากเซลล์ merge เพี้ยน
    แต่คงราคาซ้ำไว้ เช่น 57,600 / 57,600)"""
    out = []
    for c in row:
        c = clean(c)
        if c and (not out or c != out[-1] or is_price(c)):
            out.append(c)
    return out


HEADER_WORDS = ("(Satellite", "(Sattellite", "(Resolution", "Mode",
                "Single Look", "Path Image (บาท)", "New Acquisition")


def is_header(cells):
    joined = " ".join(cells)
    return any(w in joined for w in HEADER_WORDS) and not any(
        is_price(c) for c in cells
    )


# ---------------------------------------------------------------- row parser
def parse_table(table, spec, rows):
    """แปลงตาราง 1 ชุดเป็นแถวข้อมูลตาม spec:
      category, unit, satellite (None = อ่านจากคอลัมน์แรก),
      price_cols = ชื่อคอลัมน์ราคาเรียงตามลำดับที่พบ,
      has_pol = มีคอลัมน์ Polarization,
      unit_overrides = {ชื่อแถว: หน่วย}
    แถวหมายเหตุ (มี bullet) ติดกับทุกแถวข้อมูลในกลุ่มก่อนหน้า
    """
    block = []  # แถวข้อมูลที่รอรับหมายเหตุ
    for raw in table:
        has_bullet = any(BULLET in (c or "") for c in raw)
        cells = compact(raw)
        if not cells:
            continue
        # แถวชื่อกลุ่มดาวเทียมเรดาร์ เช่น "RADARSAT-2 (C band)"
        if len(cells) == 1 and re.search(r"\([CX] band\)", cells[0]):
            spec["satellite"] = cells[0]
            continue
        if has_bullet or not any(is_price(c) for c in cells):
            if is_header(cells):
                continue
            note = " | ".join(c for c in cells if c)
            for r in block:
                r["หมายเหตุ"] = (r["หมายเหตุ"] + " | " + note).strip(" |")
            continue
        # แถวข้อมูล: ชื่อ, ความละเอียด, [polarization], ราคา...
        name = cells[0]
        res = next((c for c in cells[1:] if re.search(r"(cm|m)\.?\s*$|\dx", c.replace(" ", ""))), "")
        pol = ""
        if spec.get("has_pol"):
            pol_idx = cells.index(res) + 1
            if pol_idx < len(cells) and not is_price(cells[pol_idx]):
                pol = cells[pol_idx]
        prices = [c for c in cells if is_price(c)]
        row = {c: None for c in COLUMNS}
        row["หมวด"] = spec["category"]
        row["ดาวเทียม (Satellite)"] = spec.get("satellite") or name
        row["โหมด (Mode)"] = name if spec.get("satellite") else ""
        row["รายละเอียดภาพ (Resolution)"] = res
        row["Polarization"] = pol
        row["หน่วย"] = spec.get("unit_overrides", {}).get(name, spec["unit"])
        row["หมายเหตุ"] = ""
        for col, val in zip(spec["price_cols"], prices):
            row[col] = price(val)
        # เริ่มกลุ่มใหม่เมื่อแถวก่อนหน้าปิดท้ายด้วยหมายเหตุแล้ว
        if block and block[-1]["หมายเหตุ"]:
            block = []
        block.append(row)
        rows.append(row)


CAT_VHR = "รายละเอียดสูงมาก (30–50 ซม.)"
CAT_HR = "รายละเอียดสูง (60 ซม.–2 ม.)"
CAT_MR = "รายละเอียดปานกลาง"
CAT_RADAR = "ระบบเรดาร์"
SQKM = "บาท/ตร.กม."
SCENE = "บาท/ภาพ"

# หมายเหตุกลุ่มที่เขียนเรียบเรียงใหม่ให้อ่านชัด (เนื้อหาตาม PDF)
NOTE_OVERRIDES = {
    **dict.fromkeys(
        ["Pléiades NEO", "WorldView-4", "SuperView-2", "WorldView-1",
         "WorldView-2", "WorldView-3", "GeoEye-1", "Pléiades",
         "EarthScanner", "SuperView-1", "KOMPSAT-3"],
        "ในคลัง: สั่งขั้นต่ำ 25 ตร.กม. / สั่งถ่าย: สั่งขั้นต่ำ 100 ตร.กม. / "
        "ราคาสำหรับ level Primary (PAN, MS, Pansharpened)"),
    "SKYSAT": "ในคลัง: สั่งขั้นต่ำ 1,250 ตร.กม. เข้าดูผ่าน API/Explorer / "
              "สั่งถ่าย: โปรดติดต่อเจ้าหน้าที่",
    **dict.fromkeys(
        ["QuickBird", "GaoFen-7", "Jilin", "DailyVision", "GaoFen-2", "IKONOS"],
        "ในคลัง: สั่งขั้นต่ำ 25 ตร.กม. ราคาสำหรับ level Primary (PAN, MS, "
        "Pansharpened) / สั่งถ่าย: สั่งขั้นต่ำ 100 ตร.กม."),
    **dict.fromkeys(
        ["SPOT-6", "SPOT-7"],
        "ในคลัง: สั่งขั้นต่ำ 100 ตร.กม. / สั่งถ่าย: สั่งขั้นต่ำ 500 ตร.กม."),
    "PLANETSCOPE": "ราคาในคลังแบบ Access+Download / สั่งขั้นต่ำ 100 ตร.กม. / "
                   "ระยะเวลาสัญญา 1 ปี / เข้าดูและดาวน์โหลดผ่าน Planet Explorer, "
                   "Planet API, Desktop GIS",
}


def extract():
    rows = []
    with pdfplumber.open(PDF) as pdf:
        p = [pg.extract_tables() for pg in pdf.pages]

    # หน้า 1: ออปติคัลรายละเอียดสูงมาก
    parse_table(p[0][0], dict(category=CAT_VHR, unit=SQKM,
                              price_cols=[ARCHIVE, TASKING]), rows)
    # หน้า 2: ออปติคัลรายละเอียดสูง + Video/Night
    parse_table(p[1][0], dict(category=CAT_HR, unit=SQKM,
                              price_cols=[ARCHIVE, TASKING]), rows)
    parse_table(p[1][2], dict(category=CAT_HR, unit=SQKM,
                              unit_overrides={"Video Constellation": "บาท/30 วินาที"},
                              price_cols=[ARCHIVE, TASKING]), rows)
    # หน้า 3: SPOT + ไทยโชต
    parse_table(p[2][0], dict(category=CAT_HR, unit=SQKM,
                              price_cols=[ARCHIVE, TASKING]), rows)
    parse_table(p[2][1], dict(category=CAT_HR, unit=SCENE,
                              price_cols=[ARCHIVE, TASKING]), rows)
    # หน้า 4: LANDSAT + PLANETSCOPE
    parse_table(p[3][0], dict(category=CAT_MR, unit=SCENE,
                              price_cols=[ARCHIVE, TASKING]), rows)
    parse_table(p[3][1], dict(category=CAT_MR, unit="บาท/ตร.กม./ปี",
                              price_cols=[ARCHIVE, MONITOR]), rows)
    # หน้า 5: RADARSAT-2
    parse_table(p[4][0], dict(category=CAT_RADAR, unit=SCENE, satellite=None,
                              price_cols=[SLC, PATH_IMG]), rows)
    # หน้า 6: TerraSAR-X + COSMO SkyMed
    parse_table(p[5][0], dict(category=CAT_RADAR, unit=SCENE, satellite=None,
                              price_cols=[ARCHIVE, TASKING]), rows)
    parse_table(p[5][1], dict(category=CAT_RADAR, unit=SCENE, satellite=None,
                              has_pol=True, price_cols=[NEW_ACQ]), rows)
    # หน้า 7: GaoFen-3
    parse_table(p[6][0], dict(category=CAT_RADAR, unit=SCENE, satellite=None,
                              has_pol=True, price_cols=[ARCHIVE, TASKING]), rows)

    for r in rows:
        key = r["โหมด (Mode)"] or r["ดาวเทียม (Satellite)"]
        if key in NOTE_OVERRIDES:
            r["หมายเหตุ"] = NOTE_OVERRIDES[key]

    df = pd.DataFrame(rows, columns=COLUMNS)
    for col in PRICE_COLUMNS:
        df[col] = df[col].astype("Int64")
    return df


# ---------------------------------------------------------------- exporters
def export(df):
    # CSV (UTF-8 BOM เปิดใน Excel ได้)
    df.to_csv(BASE / f"{STEM}.csv", index=False, encoding="utf-8-sig")

    # XLSX
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    xlsx = BASE / f"{STEM}.xlsx"
    with pd.ExcelWriter(xlsx, engine="openpyxl") as xw:
        df.to_excel(xw, index=False, sheet_name="Price List")
        ws = xw.sheets["Price List"]
        header_fill = PatternFill("solid", fgColor="1F4E79")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        widths = [24, 26, 26, 20, 22, 16, 16, 14, 12, 14, 14, 16, 70]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        for row in ws.iter_rows(min_row=2):
            for cell in row[5:11]:
                cell.number_format = "#,##0"
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        note_row = ws.max_row + 2
        ws.cell(note_row, 1, f"** {VAT_NOTE}").font = Font(italic=True, color="C00000")

    # HTML
    html_table = df.astype("object").fillna("").to_html(
        index=False, border=0, justify="center", classes="price",
        formatters={c: (lambda v: f"{v:,}" if v != "" else "") for c in PRICE_COLUMNS})
    html = f"""<!DOCTYPE html>
<html lang="th">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>GISTDA Price List — ราคาข้อมูลจากดาวเทียม</title>
<style>
  body {{ font-family: 'Segoe UI', 'Leelawadee UI', Tahoma, sans-serif; margin: 2rem; color: #222; }}
  h1 {{ font-size: 1.4rem; color: #1F4E79; }}
  .vat {{ color: #C00000; font-style: italic; }}
  .src {{ color: #666; font-size: .85rem; }}
  .wrap {{ overflow-x: auto; }}
  table.price {{ border-collapse: collapse; width: 100%; font-size: .9rem; }}
  table.price th {{ background: #1F4E79; color: #fff; padding: .5rem .6rem; position: sticky; top: 0; }}
  table.price td {{ border-bottom: 1px solid #ddd; padding: .4rem .6rem; }}
  table.price tr:nth-child(even) {{ background: #F2F7FB; }}
  table.price tr:hover {{ background: #DCE9F5; }}
</style>
</head>
<body>
<h1>ราคาข้อมูลจากดาวเทียม (GISTDA Price List)</h1>
<p class="vat">** {VAT_NOTE}</p>
<p class="src">ที่มา: <a href="https://www.gistda.or.th/download/Gistda_Price_List.pdf">Gistda_Price_List.pdf</a>
 — สกัดข้อมูลอัตโนมัติ | ราคาว่าง = ไม่มีบริการ (N/A)</p>
<div class="wrap">
{html_table}
</div>
</body>
</html>"""
    (BASE / f"{STEM}.html").write_text(html, encoding="utf-8")


if __name__ == "__main__":
    df = extract()
    export(df)
    print(f"rows: {len(df)}")
    print(df["หมวด"].value_counts().to_string())
    print("files:", ", ".join(f"{STEM}.{e}" for e in ("csv", "xlsx", "html")))
